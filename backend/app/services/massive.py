from __future__ import annotations

import io
import json
import os
import re
import threading
import time
import unicodedata
import uuid
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import env_float
from app.services.factiliza import FactilizaClient
from app.services.sunat_cpe import SunatCPEClient
from app.services.sunat_web import SunatWebClient

JOBS: Dict[str, Dict[str, Any]] = {}
LOCK = threading.Lock()
OUTPUT_DIR = Path(os.getenv("ROD_OUTPUT_DIR", "outputs"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

DOC_ALIASES = ("DNI RUC", "RUC DNI", "DOCUMENTO", "NRO DOCUMENTO", "NUMERO DOCUMENTO", "DNI", "RUC", "DOCUMENTO PROVEEDOR", "RUC PROVEEDOR")
NAME_ALIASES = ("NOMBRES Y APELLIDOS RAZON SOCIAL", "NOMBRE RAZON SOCIAL", "RAZON SOCIAL", "NOMBRES Y APELLIDOS", "NOMBRE COMPLETO", "APORTANTE", "PROVEEDOR", "NOMBRES")
TYPE_ALIASES = ("TIPO COMPROBANTE", "TIPO DE COMPROBANTE", "TIPO COMP", "CODIGO COMPROBANTE", "COD COMPROBANTE")
CPE_ALIASES = ("NRO COMPROBANTE", "NUMERO COMPROBANTE", "N COMPROBANTE", "COMPROBANTE")
SERIE_ALIASES = ("SERIE", "SERIE COMPROBANTE", "NRO SERIE", "NUMERO SERIE")
NUM_ALIASES = ("NUMERO", "NRO", "CORRELATIVO", "NUMERO CPE", "NRO CPE")
DATE_ALIASES = ("FECHA EMISION", "FECHA DE EMISION", "F EMISION", "FECHA")
AMOUNT_ALIASES = ("MONTO", "IMPORTE", "TOTAL", "MONTO COMPROBANTE", "IMPORTE TOTAL")
TYPE_MAP = {"1":"01","01":"01","FACTURA":"01","2":"03","03":"03","BOLETA":"03","3":"R1","R1":"R1","RH":"R1","RHE":"R1","RECIBO POR HONORARIOS":"R1"}
HEADER_HINTS = ("DNI","RUC","DOCUMENTO","NOMBRE","RAZON SOCIAL","APORTANTE","PROVEEDOR","COMPROBANTE","SERIE","FECHA","MONTO","APORTE EFECTIVO","APORTE ESPECIE")


def clean(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


def norm(v: Any) -> str:
    t = unicodedata.normalize("NFKD", clean(v).upper())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9]+", " ", t)).strip()


def similarity(a: str, b: str) -> float:
    a, b = norm(a), norm(b)
    if not a or not b: return 0.0
    if a == b: return 1.0
    sa, sb = " ".join(sorted(a.split())), " ".join(sorted(b.split()))
    return max(SequenceMatcher(None,a,b).ratio(), SequenceMatcher(None,sa,sb).ratio())


def digits(v: Any) -> str:
    s = clean(v)
    if s.endswith(".0") and s[:-2].isdigit(): s = s[:-2]
    return re.sub(r"\D", "", s)


def find_col(cols: List[str], aliases: Tuple[str, ...]) -> Optional[str]:
    lookup = {norm(c): c for c in cols}
    for a in aliases:
        if norm(a) in lookup: return lookup[norm(a)]
    for a in aliases:
        na = norm(a)
        if len(na) < 4: continue
        for nc, raw in lookup.items():
            if na in nc or nc in na: return raw
    return None


def header_score(row: pd.Series) -> int:
    joined = " | ".join(norm(v) for v in row.tolist() if clean(v))
    score = sum(1 for h in HEADER_HINTS if norm(h) in joined)
    if "DNI" in joined or "RUC" in joined: score += 2
    if "NOMBRE" in joined or "RAZON SOCIAL" in joined: score += 2
    if "COMPROBANTE" in joined: score += 2
    return score


def unique_headers(vals: List[Any]) -> List[str]:
    used: Dict[str,int] = {}; out=[]
    for i,v in enumerate(vals):
        base = clean(v) or f"COLUMNA_{i+1}"
        used[base] = used.get(base,0)+1
        out.append(base if used[base]==1 else f"{base}_{used[base]}")
    return out


def table_from_raw(raw: pd.DataFrame) -> Tuple[pd.DataFrame,int]:
    if raw.empty: return pd.DataFrame(),0
    candidates=[(i,header_score(raw.iloc[i])) for i in range(min(35,len(raw)))]
    h,best=max(candidates,key=lambda x:x[1])
    if best<3:
        h=next((i for i in range(min(35,len(raw))) if raw.iloc[i].notna().sum()>=2),0)
    df=raw.iloc[h+1:].copy(); df.columns=unique_headers(raw.iloc[h].tolist()); df=df.dropna(how="all")
    return df,h


def parse_file(content: bytes, filename: str) -> List[Dict[str,Any]]:
    rows=[]
    if filename.lower().endswith(".csv"):
        raw=pd.read_csv(io.BytesIO(content),header=None,dtype=object,sep=None,engine="python")
        df,h=table_from_raw(raw)
        for n,(_,r) in enumerate(df.iterrows(),start=h+2):
            d={str(c):r[c] for c in df.columns}; d["_ROD_HOJA"]="CSV"; d["_ROD_FILA"]=n; rows.append(d)
        return rows
    xls=pd.ExcelFile(io.BytesIO(content))
    for sheet in xls.sheet_names:
        try: raw=pd.read_excel(xls,sheet_name=sheet,header=None,dtype=object)
        except Exception: continue
        df,h=table_from_raw(raw)
        for n,(_,r) in enumerate(df.iterrows(),start=h+2):
            d={str(c):r[c] for c in df.columns}; d["_ROD_HOJA"]=str(sheet); d["_ROD_FILA"]=n; rows.append(d)
    return rows


def get_doc(row: Dict[str,Any]) -> str:
    cols=list(row.keys())
    for a in DOC_ALIASES:
        c=find_col(cols,(a,))
        if c:
            d=digits(row.get(c))
            if len(d) in {8,11}: return d
    return ""


def get_name(row: Dict[str,Any]) -> str:
    c=find_col(list(row.keys()),NAME_ALIASES)
    return clean(row.get(c)) if c else ""


def fmt_date(v: Any) -> str:
    if not clean(v): return ""
    if isinstance(v,(datetime,pd.Timestamp)): return v.strftime("%d/%m/%Y")
    try: return pd.to_datetime(v,dayfirst=True).strftime("%d/%m/%Y")
    except Exception: return clean(v)


def amount(v: Any) -> Optional[float]:
    s=clean(v).replace("S/","").replace(",","")
    try: return float(s) if s else None
    except Exception: return None


def get_cpe(row: Dict[str,Any]) -> Dict[str,Any]:
    cols=list(row.keys()); tc=find_col(cols,TYPE_ALIASES); sc=find_col(cols,SERIE_ALIASES); nc=find_col(cols,NUM_ALIASES); cc=find_col(cols,CPE_ALIASES); dc=find_col(cols,DATE_ALIASES); mc=find_col(cols,AMOUNT_ALIASES)
    tipo=TYPE_MAP.get(norm(row.get(tc)) if tc else "",""); serie=clean(row.get(sc)).upper() if sc else ""; numero=clean(row.get(nc)) if nc else ""; combined=clean(row.get(cc)) if cc else ""
    if combined and (not serie or not numero):
        m=re.match(r"^\s*([A-Za-z0-9]+)\s*[-/]\s*([A-Za-z0-9]+)\s*$",combined)
        if m: serie=serie or m.group(1).upper(); numero=numero or m.group(2)
    if not tipo and serie:
        if serie.startswith("F"): tipo="01"
        elif serie.startswith("B"): tipo="03"
        elif serie.startswith("E"): tipo="R1"
    return {"tipo":tipo,"serie":serie,"numero":re.sub(r"\.0$","",numero),"fecha":fmt_date(row.get(dc)) if dc else "","monto":amount(row.get(mc)) if mc else None,"original":combined}


def set_job(job_id: str, **kw: Any) -> None:
    with LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(kw); JOBS[job_id]["updated_at"]=time.time()


def status(job_id: str) -> Optional[Dict[str,Any]]:
    with LOCK:
        j=dict(JOBS[job_id]) if job_id in JOBS else None
    if j: j.pop("output_path",None)
    return j


def output_path(job_id: str) -> Optional[Path]:
    with LOCK: p=JOBS.get(job_id,{}).get("output_path")
    path=Path(p) if p else None
    return path if path and path.exists() else None


def write_excel(job_id: str, filename: str, sheets: Dict[str,List[Dict[str,Any]]]) -> Path:
    base=re.sub(r"[^A-Za-z0-9_-]+","_",Path(filename).stem)[:70] or "ROD"
    path=OUTPUT_DIR/f"ROD_{base}_{job_id[:8]}.xlsx"
    with pd.ExcelWriter(path,engine="openpyxl") as w:
        for name,data in sheets.items(): pd.DataFrame(data).to_excel(w,sheet_name=name[:31],index=False)
    wb=load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="173B61"); c.alignment=Alignment(vertical="center")
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(48,max(10,max(len(clean(c.value)) for c in col[:200])+2))
    if "CONTROL_INTERNO" in wb.sheetnames: wb["CONTROL_INTERNO"].sheet_state="hidden"
    wb.save(path); return path


def process(job_id: str, content: bytes, filename: str, mode: str, use_dni: bool, use_ruc: bool, use_cpe: bool, use_reps: bool, pj_queue: bool) -> None:
    fact=FactilizaClient(); cpe_client=SunatCPEClient(); sunat=None; dni_cache={}; ruc_cache={}; cpe_cache={}; reps_cache={}
    try:
        set_job(job_id,status="RUNNING",stage="Leyendo Excel",progress=3)
        rows=parse_file(content,filename)
        if not rows: raise RuntimeError("No se encontraron filas utilizables en el archivo.")
        total=len(rows); set_job(job_id,total=total,stage="Preparando consultas",progress=6)
        general=[]; dnis=[]; rucs=[]; cpes=[]; reps=[]; obs=[]; pj=[]; internal=[]
        if use_ruc or use_reps: sunat=SunatWebClient()
        for i,row in enumerate(rows,start=1):
            doc=get_doc(row); name=get_name(row); notes=[]; observed=False; technical=False
            out={"HOJA":row.get("_ROD_HOJA",""),"FILA_EXCEL":row.get("_ROD_FILA",""),"DOCUMENTO":doc,"TIPO_DOCUMENTO":"DNI" if len(doc)==8 else "RUC" if len(doc)==11 else "","NOMBRE_RAZON_DECLARADO":name,"RESULTADO_IDENTIDAD":"","NOMBRE_FACTILIZA":"","SIMILITUD_DNI":"","RESULTADO_RUC":"","RAZON_SOCIAL_SUNAT":"","ESTADO_RUC":"","CONDICION_RUC":"","DOMICILIO_FISCAL":"","SIMILITUD_RUC":"","RESULTADO_COMPROBANTE":"","OBSERVACION_ROD":"","ESTADO_FINAL":"CORRECTO"}
            if len(doc)==8 and use_dni:
                if doc not in dni_cache: set_job(job_id,stage=f"Factiliza DNI {doc}"); dni_cache[doc]=fact.consultar(doc,"")
                raw=dni_cache[doc]; official=clean(raw.get("nombre")); sim=similarity(name,official) if name and official else None; match=None if sim is None else sim>=env_float("MATCH_MIN_SIMILITUD_DNI",0.82)
                result="COINCIDE" if match is True else "NO COINCIDE" if match is False else "DNI ENCONTRADO" if raw.get("ok") else raw.get("status","ERROR")
                out.update({"RESULTADO_IDENTIDAD":result,"NOMBRE_FACTILIZA":official,"SIMILITUD_DNI":round(sim,3) if sim is not None else ""}); dnis.append({"DNI":doc,"NOMBRE_DECLARADO":name,"NOMBRE_FACTILIZA":official,"SIMILITUD":out["SIMILITUD_DNI"],"RESULTADO":result,"OK_TECNICO":raw.get("ok",False)})
                if not raw.get("ok"): technical=True; notes.append("DNI pendiente por falla/configuración de Factiliza")
                elif match is False: observed=True; notes.append("DNI y nombre no coinciden")
            elif len(doc)==11 and use_ruc:
                if doc not in ruc_cache:
                    set_job(job_id,stage=f"SUNAT RUC {doc}")
                    try: ruc_cache[doc]=sunat.consultar(doc,"")
                    except Exception as e:
                        try: sunat.close(); sunat=SunatWebClient(); ruc_cache[doc]=sunat.consultar(doc,"")
                        except Exception as e2: ruc_cache[doc]={"ok":False,"status":"ERROR_SUNAT_WEB","message":str(e2)}
                raw=ruc_cache[doc]; razon=clean(raw.get("razon_social")); sim=similarity(name,razon) if name and razon else None; ok_t=env_float("MATCH_MIN_SIMILITUD_RUC_OK",0.92); rev_t=env_float("MATCH_MIN_SIMILITUD_RUC_REVISAR",0.84)
                if not raw.get("ok"): result=raw.get("status","ERROR"); technical=True; notes.append("RUC pendiente por falla técnica de SUNAT")
                elif sim is None: result="RUC ENCONTRADO"
                elif sim>=ok_t: result="COINCIDE"
                elif sim>=rev_t: result="REVISAR SIMILITUD"; observed=True; notes.append("RUC y razón social requieren revisión")
                else: result="NO COINCIDE"; observed=True; notes.append("RUC y razón social no coinciden")
                estado=clean(raw.get("estado")); condicion=clean(raw.get("condicion"))
                if raw.get("ok") and ("BAJA" in norm(estado) or "NO HABIDO" in norm(condicion)): observed=True; notes.append("Estado/condición SUNAT requiere revisión")
                out.update({"RESULTADO_RUC":result,"RAZON_SOCIAL_SUNAT":razon,"ESTADO_RUC":estado,"CONDICION_RUC":condicion,"DOMICILIO_FISCAL":clean(raw.get("domicilio_fiscal")),"SIMILITUD_RUC":round(sim,3) if sim is not None else ""}); rucs.append({"RUC":doc,"RAZON_DECLARADA":name,"RAZON_SUNAT":razon,"SIMILITUD":out["SIMILITUD_RUC"],"RESULTADO":result,"ESTADO":estado,"CONDICION":condicion,"DOMICILIO_FISCAL":out["DOMICILIO_FISCAL"],"OK_TECNICO":raw.get("ok",False)})
                if use_reps and doc.startswith("20") and raw.get("ok"):
                    if doc not in reps_cache:
                        try: set_job(job_id,stage=f"Representantes {doc}"); reps_cache[doc]=sunat.representantes(doc,razon)
                        except Exception as e: reps_cache[doc]={"ok":False,"status":"ERROR_REPRESENTANTES","representantes":[],"message":str(e)}
                    reps.extend(reps_cache[doc].get("representantes",[]))
            cpe=get_cpe(row) if use_cpe and mode.upper() in {"4D","AUTO","CPE"} else {}
            if cpe and any(cpe.get(k) for k in ("serie","numero","original")):
                if not cpe.get("serie") or not cpe.get("numero"): observed=True; out["RESULTADO_COMPROBANTE"]="FORMATO INCOMPLETO"; notes.append("No se pudo separar serie y número del comprobante")
                elif not cpe.get("tipo") or not cpe.get("fecha"): observed=True; out["RESULTADO_COMPROBANTE"]="DATOS INCOMPLETOS"; notes.append("Falta tipo o fecha del comprobante")
                elif len(doc) not in {8,11}: observed=True; out["RESULTADO_COMPROBANTE"]="EMISOR INVÁLIDO"; notes.append("Documento emisor inválido")
                else:
                    key=json.dumps({"d":doc,**cpe},sort_keys=True,default=str)
                    if key not in cpe_cache: set_job(job_id,stage=f"SUNAT CPE {cpe.get('serie')}-{cpe.get('numero')}"); cpe_cache[key]=cpe_client.validar({"tipo":cpe.get("tipo"),"ruc":doc,"serie":cpe.get("serie"),"numero":cpe.get("numero"),"fecha":cpe.get("fecha"),"monto":cpe.get("monto")})
                    cr=cpe_cache[key]; human=clean(cr.get("human_status") or cr.get("status")); out["RESULTADO_COMPROBANTE"]=human; cpes.append({"DOCUMENTO_EMISOR":doc,"TIPO":cpe.get("tipo"),"SERIE":cpe.get("serie"),"NUMERO":cpe.get("numero"),"FECHA":cpe.get("fecha"),"MONTO":cpe.get("monto"),"RESULTADO":human,"ESTADO_CP":cr.get("estado_cp",""),"ESTADO_RUC":cr.get("estado_ruc",""),"CONDICION_RUC":cr.get("condicion_ruc",""),"OBSERVACIONES_SUNAT":clean(cr.get("observaciones")),"INTENTOS":cr.get("attempts",""),"OK_TECNICO":cr.get("ok",False)})
                    if not cr.get("ok"): technical=True; notes.append("Comprobante pendiente por falla técnica/configuración")
                    elif cr.get("estado_cp") not in {"1","3"}: observed=True; notes.append(f"Comprobante: {human}")
            if not doc: observed=True; notes.append("No se detectó DNI/RUC válido")
            out["ESTADO_FINAL"]="PENDIENTE_TECNICO" if technical else "OBSERVADO" if observed else "CORRECTO"; out["OBSERVACION_ROD"]=" | ".join(dict.fromkeys(notes))
            for k,v in row.items():
                if not str(k).startswith("_ROD_"): out[f"ORIGINAL_{k}"]=v
            general.append(out)
            if out["ESTADO_FINAL"]!="CORRECTO": obs.append(out.copy())
            if pj_queue and len(doc)==8 and out["ESTADO_FINAL"]=="CORRECTO": pj.append({"DNI":doc,"NOMBRE_VALIDADO":out.get("NOMBRE_FACTILIZA") or name,"CONDENA":"","PRISION PREVENTIVA":"","ESTADO_PJ":"PENDIENTE_ASISTIDO"})
            internal.append({"HOJA":out["HOJA"],"FILA":out["FILA_EXCEL"],"DOCUMENTO":doc,"ESTADO_FINAL":out["ESTADO_FINAL"],"OBSERVACION":out["OBSERVACION_ROD"],"FECHA_HORA":datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            set_job(job_id,processed=i,progress=min(94,8+int(i/total*86)))
        set_job(job_id,stage="Generando Excel",progress=96)
        path=write_excel(job_id,filename,{"RESULTADO_GENERAL":general,"DNI_FACTILIZA":dnis,"RUC_SUNAT":rucs,"COMPROBANTES_SUNAT":cpes,"REPRESENTANTES":reps,"OBSERVADOS":obs,"PJ_PENDIENTES":pj,"CONTROL_INTERNO":internal})
        counts={"rows":total,"correct":sum(x["ESTADO_FINAL"]=="CORRECTO" for x in general),"observed":sum(x["ESTADO_FINAL"]=="OBSERVADO" for x in general),"technical":sum(x["ESTADO_FINAL"]=="PENDIENTE_TECNICO" for x in general),"dni_unique":len(dni_cache),"ruc_unique":len(ruc_cache),"cpe_unique":len(cpe_cache),"representatives":len(reps),"pj_pending":len(pj)}
        set_job(job_id,status="DONE",stage="Terminado",progress=100,processed=total,counts=counts,output_path=str(path),download_name=path.name)
    except Exception as e:
        set_job(job_id,status="ERROR",stage="Error",error=f"{type(e).__name__}: {e}")
    finally:
        if sunat:
            try: sunat.close()
            except Exception: pass


def submit(content: bytes, filename: str, mode: str="AUTO", use_dni: bool=True, use_ruc: bool=True, use_cpe: bool=True, use_reps: bool=False, pj_queue: bool=True) -> Dict[str,Any]:
    job_id=uuid.uuid4().hex
    with LOCK: JOBS[job_id]={"job_id":job_id,"status":"QUEUED","stage":"En cola","progress":0,"processed":0,"total":0,"filename":filename,"mode":mode.upper(),"counts":{},"error":"","created_at":time.time(),"updated_at":time.time()}
    threading.Thread(target=process,args=(job_id,content,filename,mode,use_dni,use_ruc,use_cpe,use_reps,pj_queue),daemon=True).start()
    return status(job_id) or {"job_id":job_id,"status":"QUEUED"}
